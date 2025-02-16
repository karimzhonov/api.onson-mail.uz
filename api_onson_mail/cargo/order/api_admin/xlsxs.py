import openpyxl
from io import BytesIO
from django.conf import settings
import openpyxl.drawing
import openpyxl.drawing.image
from cargo.models import User
from contrib.qrcode import generate_qrcode
INPUT_PATH = settings.BASE_DIR / "cargo/order/xlsxs/invoice.xlsx"  # Укажите путь к исходному файлу


def _refactor_sheet(sheet, order, phones, start=0):
    qrcode = openpyxl.drawing.image.Image(generate_qrcode(order.id))
    qrcode.height = 220
    qrcode.width = 220
    qrcode.anchor = f"K{3+start}"
    sheet.add_image(qrcode)
    sheet.cell(row=3 + start, column=7, value=order.number)
    sheet.cell(row=16 + start, column=11, value=f'{order.parts.number}-OTPRAVKA IPAK YOLI')
    sheet.cell(row=13 + start, column=3, value=order.client.fio)
    sheet.cell(row=15 + start, column=3, value=order.client.address)
    sheet.cell(row=17 + start, column=3, value=order.client.passport)
    sheet.cell(row=17 + start, column=6, value=order.client.pnfl)
    sheet.cell(row=19 + start, column=3, value=phones)

    products = list(order.products.get("product_counts", {}).values())
    st = 22 + start
    for i, product in enumerate(products[:25]):
        print(product)
        sheet.cell(row=st, column=1, value=i+1)
        sheet.cell(row=st, column=2, value=product.get('product', ''))
        sheet.cell(row=st, column=3, value=product.get('count_product', ''))
        sheet.cell(row=st, column=6, value=product.get('price_per_product', ''))
        sheet.cell(row=st, column=7, value=product.get('total_price', ''))
        st += 1
    st = 22 + start
    for i, product in enumerate(products[25:]):
        sheet.cell(row=st, column=9, value=i+1)
        sheet.cell(row=st, column=10, value=product.get('name', ''))
        sheet.cell(row=st, column=11, value=product.get('count_product', ''))
        sheet.cell(row=st, column=14, value=product.get('price_per_product', ''))
        sheet.cell(row=st, column=15, value=product.get('total_price', ''))
        st += 1
        
    sheet.cell(row=48 + start, column=14, value=f'{round(order.weight or 0, 1)} кг')
    sheet.cell(row=48 + start, column=5, value=str(order.create_time.date().strftime('%d.%m.%Y')))
    sheet.cell(row=49 + start, column=14, value=f'{round(order.facture_price or 0)} $')


def generate_invoice(order):
    phones = ", ".join([str(p) for p in list(User.objects.filter(clients__in=[order.client]).values_list("user__phone", flat=True))])
    wb = openpyxl.load_workbook(INPUT_PATH)
    sheet = wb.active

    _refactor_sheet(sheet, order, phones)
    _refactor_sheet(sheet, order, phones, start=51)

    io = BytesIO()
    wb.save(io)
    io.seek(0)
    return io.getvalue()
